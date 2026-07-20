"""
Reusable verification harness for openpyxl xlsx edits.

Fill the GT list with the user's SOURCE data (book name + price, None = leave blank).
Run against the SAVED output file. Reports PASS/FAIL with per-item diffs —
works WITHOUT a spreadsheet engine (no LibreOffice needed).

Usage:
  guix shell python python-openpyxl -- python3 verify_xlsx_ground_truth.py
"""
import openpyxl
from openpyxl.utils import get_column_letter

OUT = "/home/brokenshine/Downloads/智能2543班级应缴纳教材费用明细表（2026下半）_已补录.xlsx"

# ---- FILL THIS with the user's original data (name, price) ----
# price=None means the cell should be left EMPTY (e.g. missing from a truncated image)
GT = [
    ("全新版大学高阶英语：综合训练3", 26.07),
    ("全新版大学进阶英语：综合教程（第二版）AI增强版3学生用书", 54.51),
    # ... (copy ALL items from the user's source verbatim) ...
    ("内部控制学 (第二版)", 31.60),
]

SUFFIXES = ["（打折后价格）", "(打折后价格)", "（打折后价格)", "(打折后价格）"]

def strip_suffix(h):
    if h is None:
        return None
    h = str(h)
    for s in SUFFIXES:
        if h.endswith(s):
            h = h[:-len(s)]
    h = h.strip()
    if h.startswith("11 "):   # template "11 bookname" prefix
        h = h[3:].strip()
    return h.strip()

wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]
N = len(GT)
assert N == 44, f"set GT length to {N} to match actual item count"

file_names, file_prices = [], []
for col in range(3, 3 + N):
    file_names.append(strip_suffix(ws.cell(row=2, column=col).value))
    file_prices.append(ws.cell(row=3, column=col).value)

errors = []
if len(file_names) != N:
    errors.append(f"header count={len(file_names)} expected={N}")
for i, (gn, gp) in enumerate(GT):
    if file_names[i] != gn:
        errors.append(f"[{i+1}] name mismatch\n   file: {file_names[i]!r}\n   GT  : {gn!r}")
    fp, gp_ = file_prices[i], gp
    if gp_ is None:
        if fp is not None:
            errors.append(f"[{i+1}] {gn}: expected EMPTY, got {fp!r}")
    else:
        if fp is None or abs(float(fp) - float(gp_)) > 0.001:
            errors.append(f"[{i+1}] {gn}: file={fp} expected={gp_}")

# formula range spot-checks (edit the letters to match your layout)
print("AU4 :", ws.cell(row=4, column=47).value.text if hasattr(ws.cell(row=4, column=47).value, "text") else ws.cell(row=4, column=47).value)
print("AV65:", ws.cell(row=65, column=48).value)

print("\n" + ("✅ ALL CHECKS PASS — zero misalignment, zero typo" if not errors else "❌ ERRORS:"))
for e in errors:
    print("  -", e)
